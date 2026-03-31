from __future__ import annotations
from PySide6.QtCore import QCoreApplication
from typing import TYPE_CHECKING, Type
import qtawesome as qta
from bsdd_gui.module.excel import constants
import json
import openpyxl
from bsdd_json.utils import class_utils

if TYPE_CHECKING:
    from bsdd_gui import tool
    from bsdd_gui.module.excel import ui
    from bsdd_gui.tool.property_picker import PsetDict
    from bsdd_gui.tool.excel import SettingsDict

def connect_to_main_window(
    excel: Type[tool.Excel],
    main_window: Type[tool.MainWindowWidget],
    project: Type[tool.Project],
):
    # Action uses the WidgetTool request to allow trigger routing

    action = main_window.add_action(
        "menuFile",
        "Export Excel",
        lambda: excel.request_widget(
            project.get(),
            main_window.get(),
        ),
        qta.icon("mdi6.file-excel"),

    )
    excel.set_action(main_window.get(), "open_window", action)


def connect_signals(excel: Type[tool.Excel]):
    excel.connect_internal_signals()


def retranslate_ui(excel: Type[tool.Excel],main_window:type[tool.MainWindowWidget]):
    action = excel.get_action(main_window.get(), "open_window")
    action.setText(QCoreApplication.translate("Excel", "Export Excel"))



def create_widget(data, parent, excel: Type[tool.Excel]):
    excel.show_widget(data, parent)


def register_widget(widget: ui.Widget, excel: Type[tool.Excel]):
    excel.register_widget(widget)
    widget.fw_output.file_format = constants.EXCEL_FILETYPE
    widget.fw_output.section = "paths"
    widget.fw_output.option = "excel"
    widget.fw_output.title = "get Excel-Export Path"

    widget.pb_import.setIcon(qta.icon("mdi6.tray-arrow-up"))
    widget.pb_export.setIcon(qta.icon("mdi6.tray-arrow-down"))
    widget.fw_output.load_path()


def register_fields(widget: ui.Widget, excel: Type[tool.Excel]):
    pass
    #excel.register_basic_field(widget, widget.le_name, "Name")


def register_validators(widget: ui.Widget, excel: Type[tool.Excel], util: Type[tool.Util]):
    pass
    # excel.add_validator(
    #     widget,
    #     widget.le_code,
    #     lambda v, w,: excel.is_code_valid(v, w),
    #     lambda w, v: util.set_invalid(w, not v),
    # )


def connect_widget(widget: ui.Widget, excel: Type[tool.Excel]):
    excel.connect_widget_signals(widget)

def export_settings(
    widget: ui.Widget,
    widget_tool: Type[tool.Excel],
    pp_class_view: Type[tool.PPClassView],
    pp_property_view: Type[tool.PPPropertyView],
    appdata: Type[tool.Appdata],
    popups: Type[tool.Popups],
):
    # Create Dict
    class_tree = widget.property_picker.tv_classes
    property_tree = widget.property_picker.tv_properties
    class_dict: dict[str, bool] = pp_class_view.get_check_dict(class_tree)
    property_dict: PsetDict = pp_property_view.get_check_dict(property_tree)
    settings_dict: SettingsDict = widget_tool.get_settings(widget)
    full_dict: SettingsDict = {
        "class_settings": class_dict,
        "property_settings": property_dict,
        "settings": settings_dict,
    }

    # Set Path
    text = QCoreApplication.translate("Excel", "Export Excel settings")
    old_path = appdata.get_path(constants.APPDATA_OPTION)
    new_path = popups.get_save_path(constants.SETTINGS_FILETYPE, widget.window(), old_path, text)
    if not new_path:
        return
    appdata.set_path(constants.APPDATA_OPTION, new_path)

    # Write Json
    with open(new_path, "w") as file:
        json.dump(full_dict, file)


def import_settings(
    widget: ui.Widget,
    widget_tool: Type[tool.Excel],
    pp_class_view: Type[tool.PPClassView],
    pp_property_view: Type[tool.PPPropertyView],
    appdata: Type[tool.Appdata],
    popups: Type[tool.Popups],
):
    # Handle Path
    old_path = appdata.get_path(constants.APPDATA_OPTION)
    text = QCoreApplication.translate("Excel", "Import Excel settings")
    new_path = popups.get_open_path(constants.SETTINGS_FILETYPE, widget.window(), old_path, text)
    if not new_path:
        return
    appdata.set_path(constants.APPDATA_OPTION, new_path)

    # Read Settings
    with open(new_path, "r") as file:
        full_dict: SettingsDict = json.load(file)
    class_dict = full_dict.get("class_settings", {})
    property_dict = full_dict.get("property_settings", {})
    settings_dict = full_dict.get("settings", {})

    # Fill Fields and Checkstates
    class_tree = widget.property_picker.tv_classes
    property_tree = widget.property_picker.tv_properties
    pp_class_view.set_check_dict(class_dict, class_tree)
    pp_property_view.set_check_dict(property_dict, property_tree)
    widget_tool.set_settings(widget, settings_dict)
    pass

def export_excel(    widget: ui.Widget,
    excel: Type[tool.Excel],
    pp_class_view: Type[tool.PPClassView],
    pp_property_view: Type[tool.PPPropertyView],
    appdata: Type[tool.Appdata],
    popups: Type[tool.Popups],
    util:Type[tool.Util],
    project:type[tool.Project]):
    
    excel.sync_to_model(widget, widget.bsdd_data)
    class_settings = pp_class_view.get_check_dict(widget.property_picker.tv_classes)
    property_settings = pp_property_view.get_check_dict(widget.property_picker.tv_properties)
    base_settings = excel.get_settings(widget)
    title = QCoreApplication.translate("Excel", "Export Excel")
    
    # waiting_worker, waiting_thread, waiting_widget = util.create_waiting_widget(title)
    bsdd_dict = project.get()

    bsdd_classes = [c for c in bsdd_dict.Classes if  c.ClassType == "Class" and class_settings.get(c.Code,True)]
    root_classes = [c for c in bsdd_classes if not c.ParentClassCode ]
    psets = [c for c in bsdd_dict.Classes if  c.ClassType == "GroupOfProperties"]
    COLUMN_COUNT = 6
    MAX_ENTRIES = 8

    workbook = openpyxl.Workbook()
    overview_workbook = workbook.active
    overview_workbook.title = QCoreApplication.translate("Excel","Overview")
    overview_row = 2
    overview_workbook.cell(1,1,"Code")
    overview_workbook.cell(1,2,"Name")
    overview_workbook.cell(1,3,"Definition")


    for bsdd_class in root_classes:
        overview_workbook.cell(overview_row,1,bsdd_class.Code)
        overview_workbook.cell(overview_row,2,bsdd_class.Name)
        overview_workbook.cell(overview_row,3,bsdd_class.Definition)
        overview_row+=1
        property_dict = property_settings.get(bsdd_class.Code,{})
        overview_workbook = workbook.create_sheet(bsdd_class.Name)
        row, column = 1,1
        max_row = 0
        max_row = excel.draw_class(bsdd_class,row,column,overview_workbook,property_dict,bsdd_dict)
        items_in_row = 1
        column +=COLUMN_COUNT
        for child in excel.get_all_children(bsdd_class,bsdd_dict):
            if not class_settings.get(child.Code,True):
                continue


            overview_workbook.cell(overview_row,1,bsdd_class.Code)
            overview_workbook.cell(overview_row,2,bsdd_class.Name)
            overview_workbook.cell(overview_row,3,bsdd_class.Definition)
            overview_row+=1

            property_dict = property_settings.get(child.Code,{})

            bottom_row = excel.draw_class(child,row,column,overview_workbook,property_dict,bsdd_dict)
            column +=COLUMN_COUNT
            max_row = bottom_row if bottom_row > max_row else max_row
            items_in_row +=1

            if items_in_row > MAX_ENTRIES:
                column = 1
                row = max_row+2
                items_in_row = 0


    workbook.save(r"C:\Users\melluehc\Desktop\TestExcel\test.xlsx")


